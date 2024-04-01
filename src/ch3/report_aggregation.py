import os
import openpyxl as xl
import datetime

# ファイル名などを指定 --- (*1)
SCRIPT_DIR = os.path.dirname(__file__)
REPORT_TEMPLATE = os.path.join(SCRIPT_DIR, 'report_all_template.xlsx')
TARGET_DIR = os.path.join(SCRIPT_DIR, 'report_2024_10')
SAVE_FILE = os.path.join(SCRIPT_DIR, 'report_all_2024_10.xlsx')

# 売上を集計する関数を定義 --- (*2)
def make_report(target_dir, save_file):
    # テンプレートのExcelブックを読む --- (*3)
    template_book = xl.load_workbook(REPORT_TEMPLATE)
    template_sheet = template_book.active
    # 集計日を入力 --- (*4)
    template_sheet["B3"].value = datetime.datetime.now().strftime("%Y/%m/%d")
    # ファイル一覧を取得 --- (*5)
    files = os.listdir(target_dir)
    total = 0
    for i, file in enumerate(sorted(files)):
        # Excelブックだけを対象とする --- (*6)
        if file.endswith(".xlsx") == False:
            continue
        # 部署のExcelブックを読む --- (*7)
        full_path = os.path.join(target_dir, file)
        # 集計結果を読むためにdata_only=Trueにする
        book = xl.load_workbook(full_path, data_only=True)
        sheet = book.active
        # 必要な項目を取得 --- (*8)
        department = sheet["B3"].value
        date = sheet["B4"].value
        depratment_total = sheet["B5"].value
        print(department, date, depratment_total)
        # テンプレートに書き込む --- (*9)
        row = 8 + i - 1
        template_sheet.cell(row, 1).value = date
        template_sheet.cell(row, 2).value = department
        template_sheet.cell(row, 3).value = depratment_total
        total += depratment_total
    # 集計した結果を保存 --- (*10)
    template_book.save(save_file)
    template_book.close()
    print("合計金額:", total)

if __name__ == "__main__":
    make_report(TARGET_DIR, SAVE_FILE)
