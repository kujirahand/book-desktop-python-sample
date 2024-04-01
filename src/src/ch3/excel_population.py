import PySimpleGUI as sg
# import TkEasyGUI as sg
import openpyxl as xl

# Excelファイルから人口統計を読み込む関数
def read_population():
    # Excelファイルを読み込む --- (*1)
    EXCEL_FILE = "./population_jp.xlsx"
    workbook = xl.load_workbook(EXCEL_FILE)
    # ブックの中のシート「A」を取得 --- (*2)
    sheet = workbook["A"]
    # 都道府県ごとの総人口を得る --- (*3)
    # (メモ) 都道府県名は「I14」から「I60」のセルに入っている
    #       総人口は「L14」から「L60」のセルに入っている
    result = []
    # セル名から列番号を得る --- (*4)
    name_col_index = xl.utils.column_index_from_string("I")
    pop_col_index = xl.utils.column_index_from_string("L")
    # 連続でセルの値を取得する(14行から60行まで) --- (*5)
    for row in range(14, 60+1):
        name = sheet.cell(row, name_col_index).value
        pop = sheet.cell(row, pop_col_index).value
        man = sheet.cell(row, pop_col_index+2).value
        woman = sheet.cell(row, pop_col_index+4).value
        result.append([name, pop, man, woman])
    workbook.close()
    return result

# テーブルに結果を表示する関数 --- (*6)
def show_table(data):
    layout = [
        [sg.Table(values=data,
            headings=["都道府県", "総人口(万人)", "男性", "女性"],
            auto_size_columns=True,
            expand_x=True, expand_y=True,
            justification="right",
            font=("Arial", 14))],
        [sg.Button("閉じる")]]
    # ウィンドウを作成
    window = sg.Window("人口統計", layout, size=(600, 400))
    while True: # イベントループ
        event, _ = window.read()
        if event in ["閉じる", sg.WIN_CLOSED]:
            break
    window.close()

if __name__ == "__main__":
    # メイン処理
    data = read_population() # 人口統計を読み込む
    show_table(data) # テーブルに表示する
